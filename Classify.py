#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
☆*°☆*°(∩^o^)~━━  2017/11/23 20:37        
      (ˉ▽￣～) ~~ 一捆好葱 (*˙︶˙*)☆*°
      Fuction：Used for establishing a best classification model （Problem 3）√ ━━━━━☆*°☆*°
"""
from itertools import combinations
import time
from sklearn import metrics
from sklearn.naive_bayes import GaussianNB
from sklearn.naive_bayes import MultinomialNB
from sklearn.naive_bayes import BernoulliNB
from sklearn.neighbors import KNeighborsClassifier
from sklearn.linear_model import Lasso
from sklearn.linear_model import LogisticRegression
from sklearn.ensemble import RandomForestClassifier
from sklearn import tree
from sklearn.metrics import accuracy_score
from openpyxl import load_workbook
import cPickle as pickle
all_info = []
class_dic = {}
com_dic = {}
max_corre = 0
max_com = []
model_reuslt = []
names = ['Age', 'Sleep quality', 'Sleep latency', 'Sleep time', 'Sleep efficiency', 'Sleep disorder',
         'Daytime dyfunction']

# used for reading the excel and saving data with struture of pickle
def read_data():
	global all_info
	wb = load_workbook("..\\2017APMCM Problems\Problem A\Annex II.xlsx")
	ws = wb.get_sheet_by_name("Sheet1")
	for line in ws["A1":"L6350"]:
		values = []
		for i in range(12):
			values.append(line[i].value)
		all_info.append(values)
		# print values
	save_file = open('Annex_II.pkl', 'wb')
	pickle.dump(all_info, save_file)
	save_file.close()

# used for pre-processing the data, including extracting the needed features
def data_pre_process():
	global all_info
	save_file = open('Annex_II.pkl', 'rb')
	all_info = pickle.load(save_file)
	save_file.close()
	f1 = open('Annex_II_dataset.pkl', 'wb')
	data_set = []
	for line in all_info[1:]:
		atr = []
		# atr.append(line[1])
		# atr.append(line[2])
		for i in range(5, 12):
			atr.append(line[i])
		atr.append(line[4])
		data_set.append(atr)
	pickle.dump(data_set, f1)
	# print data_set
	f1.close()

# used for getting the training set X and labels(the kind of disease) y
def get_dataset():
	global all_info, class_dic
	save_file = open('Annex_II_dataset_all.pkl', 'rb')
	all_info = pickle.load(save_file)
	save_file.close()
	atr_datas = []
	label_data = []
	for line in all_info:
		atr_data = []
		atr_data.append(line[0])
		atr_data.extend([line[i] for i in range(2, 7)])
		atr_data.append(line[8])
		atr_datas.append(atr_data)
		if line[9] not in class_dic:
			class_dic[line[9]] = len(class_dic)
		label_data.append(class_dic[line[9]])
	print 'class_num：' + str(len(class_dic))
	return atr_datas, label_data

# used for calculation of the combination of 3 - 7 features（problem 3）
def atr_combination_save(atr_data):
	global com_dic
	for j in range(3, 8):
		c_list = list(combinations([num for num in range(len(atr_data))], j))
		com_dic[j] = []
		for ii in range(len(c_list)):
			c_list[ii] = list(c_list[ii])
		com_dic[j] = c_list

# used for extracting the particular combination of features of training set
def extract_parti_num_features(X, c_list):
	new_X = []
	for line in X:
		new_atr = []
		for i in c_list:
			new_atr.append(line[i])
		new_X.append(new_atr)
	return new_X

# used for using 7 different classification to classify, including
def any_classify_method(X, y):
	global max_corre, max_com, model_reuslt
	model = GaussianNB()
	start_time = time.time()
	model.fit(X, y)
	print('training took %fs!' % (time.time() - start_time))
	expected = y
	predicted = model.predict(X)
	corre = accuracy_score(expected, predicted)
	if max_corre < corre:
		print("\n调用scikit的朴素贝叶斯算法包GaussianNB ")
		print(model)
		print max_corre, max_com
		print '\n'
		model_reuslt.append([max_com, max_corre, 'GaussNB'])

	model = MultinomialNB(alpha=1)
	start_time = time.time()
	model.fit(X, y)
	print('training took %fs!' % (time.time() - start_time))
	expected = y
	predicted = model.predict(X)
	corre = accuracy_score(expected, predicted)
	if max_corre < corre:
		max_corre = corre
		print("\n调用scikit的朴素贝叶斯算法包MultinomialNB ")
		print(model)
		print max_corre, max_com
		print '\n'
		model_reuslt.append([max_com, max_corre, 'MultNB'])
	# print(metrics.classification_report(expected, predicted))
	# print(metrics.confusion_matrix(expected, predicted))

	model = BernoulliNB(alpha=1, binarize=0.0)
	start_time = time.time()
	model.fit(X, y)
	print('training took %fs!' % (time.time() - start_time))
	expected = y
	predicted = model.predict(X)
	corre = accuracy_score(expected, predicted)
	if max_corre < corre:
		max_corre = corre
		print("\n调用scikit的朴素贝叶斯算法包BernoulliNB ")
		print(model)
		print max_corre, max_com
		print '\n'
		model_reuslt.append([max_com, max_corre, 'BernNB'])
	# print(metrics.classification_report(expected, predicted))
	# print(metrics.confusion_matrix(expected, predicted))

	model = KNeighborsClassifier()
	start_time = time.time()
	model.fit(X, y)
	print('training took %fs!' % (time.time() - start_time))
	expected = y
	predicted = model.predict(X)
	corre = accuracy_score(expected, predicted)
	if max_corre < corre:
		max_corre = corre
		print("\n调用scikit的KNeighborsClassifier ")
		print(model)
		print max_corre, max_com
		print '\n'
		model_reuslt.append([max_com, max_corre, 'KNN'])
	# print(metrics.classification_report(expected, predicted))
	# print(metrics.confusion_matrix(expected, predicted))

	model = LogisticRegression(penalty='l2')
	start_time = time.time()
	model.fit(X, y)
	print('training took %fs!' % (time.time() - start_time))
	expected = y
	predicted = model.predict(X)
	corre = accuracy_score(expected, predicted)
	if max_corre < corre:
		max_corre = corre
		print("\n调用scikit的LogisticRegression(penalty='l2') ")
		print(model)
		print max_corre, max_com
		print
		model_reuslt.append([max_com, max_corre, 'LR'])
	# print(metrics.classification_report(expected, predicted))
	# print(metrics.confusion_matrix(expected, predicted))

	model = RandomForestClassifier(n_estimators=8)
	start_time = time.time()
	model.fit(X, y)
	print('training took %fs!' % (time.time() - start_time))
	expected = y
	predicted = model.predict(X)
	corre = accuracy_score(expected, predicted)
	if max_corre < corre:
		max_corre = corre
		print("\n调用scikit的RandomForestClassifier(n_estimators=8)  ")
		print(model)
		print max_corre, max_com
		print '\n'
		model_reuslt.append([max_com, max_corre, 'RF'])
	# print(metrics.classification_report(expected, predicted))
	# print(metrics.confusion_matrix(expected, predicted))

	model = tree.DecisionTreeClassifier()
	start_time = time.time()
	model.fit(X, y)
	print('training took %fs!' % (time.time() - start_time))
	expected = y
	predicted = model.predict(X)
	corre = accuracy_score(expected, predicted)
	if max_corre < corre:
		max_corre = corre
		print("\n调用scikit的tree.DecisionTreeClassifier() ")
		print(model)
		print max_corre, max_com
		print '\n'
		model_reuslt.append([max_com, max_corre, 'DT'])
	print(metrics.classification_report(expected, predicted))
	print(metrics.confusion_matrix(expected, predicted))


	model = Lasso(alpha=0.3)
	start_time = time.time()
	model.fit(X, y)
	print('training took %fs!' % (time.time() - start_time))
	expected = y
	predicted = model.predict(X)
	# print predicted
	for i in range(len(predicted)):
		predicted[i] = round(predicted[i])
	corre = accuracy_score(expected, predicted)
	if max_corre < corre:
		max_corre = corre
		print("\n调用scikit的Lasso ")
		print(model)
		print max_corre, max_com
		print '\n'
		model_reuslt.append([max_com, max_corre, 'Lasso'])
	# print(metrics.classification_report(expected, predicted))
	# print(metrics.confusion_matrix(expected, predicted))

if __name__ == '__main__':
	# read_data()
	# data_pre_process()
	X, y = get_dataset()
	# any_classify_method(X, y)
	atr_combination_save([i for i in range(7)])
	for key, value in com_dic.items():
		# print key, value
		for com in value:
			new_X = extract_parti_num_features(X, com)
			max_com = com
			any_classify_method(new_X, y)
	f = open('model_result.pkl', 'wb')
	pickle.dump(model_reuslt, f)
	f.close()